from __future__ import annotations

import openpyxl


def load_roles_context(path: str) -> str:
    """
    Load Roles.xlsx and return a compact text block for LLM context.
    Used to infer Responsibility Owner (Q&A) and Owner (Enablement)
    when the source text doesn't explicitly name one.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""

    lines = ["AVAILABLE TEAM MEMBERS FOR OWNER ASSIGNMENT:"]

    if "Oracle Team" in wb.sheetnames:
        lines.append("\nOracle Team (name | role | email):")
        ws = wb["Oracle Team"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            name, email, role = (str(c).strip() if c else "" for c in (row + (None, None, None))[:3])
            if name:
                lines.append(f"  {name} | {role} | {email}")

    if "Fidelity Team" in wb.sheetnames:
        lines.append("\nFidelity Team (BU/Domain | Leadership | Point of Contact):")
        ws = wb["Fidelity Team"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            bu, leadership, poc = (str(c).strip() if c else "" for c in (row + (None, None, None))[:3])
            if bu:
                lines.append(f"  {bu} | Leadership: {leadership} | POC: {poc}")

    wb.close()
    return "\n".join(lines)
