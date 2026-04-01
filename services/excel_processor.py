from __future__ import annotations

import json
import uuid
import datetime as _dt
from typing import Dict, List

import openpyxl
from openai import OpenAI


EXCEL_MAPPING_PROMPT = """
Map rows from an uploaded Excel file to the Fidelity FMR Technical Session Tracker.

TARGET SHEETS:
- ODB@AWS: Date, Issue, Status (Not Resolved|Partially Resolved|Resolved), Priority (Low|Medium|High), Oracle Tracking Request, Type (Technical Issue|Enhancement), Product, Contact, Description, Fidelity Comments, Oracle Progress/Comments
- Q&A: Question Topics, Use Case Clarification, Possible Workaround or Recommendation, Google Documentation Link, Oracle Documentation Link, Responsibility Owner, Next Step, Comment, Internal SR#
- Enablement: On-Site Date, Requester, Follow Up Enablement Topics, Owner, Status, Scheduled Date, Business Units Invloved, Internal SR#, Customer SR#
- ODB@Azure: same columns as ODB@AWS

ROUTING:
- Confirmed active SR → ODB@AWS (Azure-specific → ODB@Azure)
- Question/clarification, no SR → Q&A
- Session/workshop/training request → Enablement

Map by semantic meaning. Preserve original text. Empty source = empty string.

Return JSON: {"recommendations": [{"target_sheet": "...", "summary": "...", "confidence": "high|medium|low", "reason": "...", "row_values": {}, "source_excerpt": "..."}]}
""".strip()


def process_excel_file(file_path: str, client: OpenAI, model: str, workbook_context: Dict) -> List[Dict]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_recommendations = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(rows[0])]
        row_dicts = []

        for row in rows[1:]:
            rd = {}
            for i, v in enumerate(row):
                if i >= len(headers):
                    break
                if v is None:
                    rd[headers[i]] = ""
                elif isinstance(v, (_dt.datetime, _dt.date)):
                    rd[headers[i]] = f"{v.month}/{v.day}/{str(v.year)[2:]}"
                else:
                    rd[headers[i]] = str(v).strip()
            if any(rd.values()):
                row_dicts.append(rd)

        # Batch in groups of 20
        for i in range(0, len(row_dicts), 20):
            batch = row_dicts[i:i + 20]
            try:
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": EXCEL_MAPPING_PROMPT},
                        {"role": "user", "content": json.dumps({
                            "source_sheet": sheet_name,
                            "columns": headers,
                            "rows": batch,
                        }, default=str)},
                    ],
                    response_format={"type": "json_object"},
                )
                recs = json.loads(response.choices[0].message.content).get("recommendations", [])
                for item in recs:
                    all_recommendations.append({
                        "id": str(uuid.uuid4()),
                        "type": "addition",
                        "target_sheet": item.get("target_sheet") or "Q&A",
                        "summary": item.get("summary", "Imported from Excel"),
                        "confidence": item.get("confidence", "medium"),
                        "reason": item.get("reason", f"Mapped from Excel sheet '{sheet_name}'."),
                        "row_values": {k: (v or "") for k, v in item.get("row_values", {}).items()},
                        "source_excerpt": (item.get("source_excerpt") or "")[:1200],
                    })
            except Exception as e:
                import traceback
                print(f"ERROR in excel batch {i}: {e}")
                traceback.print_exc()

    wb.close()
    return all_recommendations