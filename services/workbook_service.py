from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl


class WorkbookService:
    def __init__(self, workbook_path: str):
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

    def _load(self):
        return openpyxl.load_workbook(self.workbook_path)

    def sheet_summaries(self) -> List[Dict[str, int]]:
        wb = self._load()
        return [
            {"name": ws.title, "rows": ws.max_row, "columns": ws.max_column}
            for ws in wb.worksheets
        ]

    def workbook_context(self) -> Dict[str, object]:
        wb = self._load()
        sheet_context = {}
        all_rows = {}

        for ws in wb.worksheets:
            headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            normalized_headers = [str(h).strip() for h in headers if h]
            examples = []
            rows_data = []

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not any(v is not None and str(v).strip() for v in row):
                    continue
                row_dict = {
                    headers[i]: self._json_safe(v)
                    for i, v in enumerate(row)
                    if i < len(headers) and headers[i]
                }
                rows_data.append(row_dict)
                if len(examples) < 6:
                    examples.append([self._json_safe(v) for v in row[:min(ws.max_column, 10)]])

            sheet_context[ws.title] = {
                "headers": normalized_headers,
                "row_count": len(rows_data),
                "sample_rows": examples,
            }
            all_rows[ws.title] = rows_data

        return {"sheets": sheet_context, "all_rows": all_rows}

    def apply_additions(self, additions: List[Dict[str, object]], output_dir: str) -> str:
        source_wb = self._load()
        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)

        for src_ws in source_wb.worksheets:
            dst_ws = out_wb.create_sheet(src_ws.title)
            self._copy_sheet(src_ws, dst_ws)

        for item in additions:
            sheet_name = item.get("target_sheet") or item.get("existing_sheet")
            if not sheet_name or sheet_name not in out_wb.sheetnames:
                continue

            ws       = out_wb[sheet_name]
            row_type = item.get("type", "addition")

            if row_type == "update":
                row_num = self._find_data_row(ws, item.get("existing_row_index", 0))
                if row_num:
                    self._write_row_values(ws, row_num, item.get("row_values", {}))
            else:
                row_num = ws.max_row + 1
                self._copy_row_style(ws, row_num)
                self._write_row_values(ws, row_num, item.get("row_values", {}))

        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(output_dir) / f"Fidelity_FMR_Technical_Session_Tracker_UPDATED_{timestamp}.xlsx"
        out_wb.save(output_path)
        return str(output_path)

    def _find_data_row(self, ws, row_index: int) -> int:
        target_row = row_index + 2
        return target_row if 2 <= target_row <= ws.max_row else None

    def _copy_sheet(self, src_ws, dst_ws):
        for row in src_ws.iter_rows():
            for cell in row:
                target = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:   target._style      = copy(cell._style)
                if cell.font:        target.font        = copy(cell.font)
                if cell.fill:        target.fill        = copy(cell.fill)
                if cell.border:      target.border      = copy(cell.border)
                if cell.alignment:   target.alignment   = copy(cell.alignment)
                if cell.protection:  target.protection  = copy(cell.protection)
                if cell.number_format: target.number_format = cell.number_format

        for key, value in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[key].width  = value.width
            dst_ws.column_dimensions[key].hidden = value.hidden

        for idx, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[idx].height = dim.height
            dst_ws.row_dimensions[idx].hidden = dim.hidden

        for merged in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged))

        dst_ws.freeze_panes = src_ws.freeze_panes
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines

    def _copy_row_style(self, ws, row_num: int):
        source_row = max(2, row_num - 1)
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col_idx)
            dst = ws.cell(row=row_num,    column=col_idx)
            if src.has_style:     dst._style       = copy(src._style)
            if src.alignment:     dst.alignment    = copy(src.alignment)
            if src.number_format: dst.number_format = src.number_format

    def _write_row_values(self, ws, row_num: int, row_values: Dict[str, object]):
        headers    = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        header_map = {self._normalize(h): idx + 1 for idx, h in enumerate(headers) if h}

        for header, value in row_values.items():
            col = header_map.get(self._normalize(header))
            if col:
                ws.cell(row=row_num, column=col, value=value)

    @staticmethod
    def _normalize(value) -> str:
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value).strip().lower())

    @staticmethod
    def _json_safe(value):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return value